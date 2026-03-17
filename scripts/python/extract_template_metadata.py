#!/usr/bin/env python3
"""Extract template metadata from Excel → JSON (for templateMetadata.ts)"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

TEMPLATE = Path(__file__).resolve().parents[2] / "template" / "DIMOP126B végleges, kalkulátor_v1.xlsx"


def main():
    wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
    hw = []
    ws = wb["Hardver"]
    hw_keys = ["hw_szamitogep", "hw_monitor", "hw_laptop", "hw_nas", "hw_router", "hw_mobil", "hw_tablet", "hw_nyomtato"]
    for i, r in enumerate(range(3, 11)):
        name = ws.cell(r, 2).value
        unit = ws.cell(r, 3).value or "db"
        netto = ws.cell(r, 5).value
        brutto = ws.cell(r, 6).value
        hw.append({
            "key": hw_keys[i] if i < len(hw_keys) else f"hw_{r}",
            "row": r,
            "name": str(name or ""),
            "unit": str(unit),
            "unitCostNet": int(netto) if netto else 0,
            "unitCostGross": int(brutto) if brutto else 0,
        })

    sw_groups = []
    ws = wb["Szoftver"]
    cur_goal = None
    cur_goal_row = None
    cur_items = []
    for r in range(5, 45):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        h = ws.cell(r, 8).value
        i = ws.cell(r, 9).value
        j = ws.cell(r, 10).value
        k = ws.cell(r, 11).value
        d = ws.cell(r, 4).value
        if b and str(b).strip().replace(".", "").isdigit():
            if cur_goal is not None and cur_items:
                sw_groups.append({"goal": cur_goal, "goalRow": cur_goal_row, "items": cur_items})
            cur_goal = str(c)[:70] if c else ""
            cur_goal_row = r
            cur_items = []
        label = str(c or h or "")[:80]
        cur_items.append({
            "row": r,
            "label": label,
            "unit": str(i or "")[:30],
            "unitCostNet": int(j) if j else 0,
            "unitCostGross": int(k) if k else 0,
        })
    if cur_goal is not None and cur_items:
        sw_groups.append({"goal": cur_goal, "goalRow": cur_goal_row, "items": cur_items})

    print(json.dumps({"hardwareItems": hw, "softwareGoalGroups": sw_groups}, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
