#!/usr/bin/env python3
"""
DIMOP126 Workbook Service
- Copy template to company folder
- Write cell values
- Trigger LibreOffice recalc (optional)
"""
import json
import argparse
import shutil
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Run: pip install openpyxl")
    exit(1)


def _cell_ref_to_tuple(cell_ref: str) -> tuple[int, int]:
    """Convert 'H14' -> (14, 8)"""
    col_letter = "".join(c for c in cell_ref if c.isalpha())
    row = int("".join(c for c in cell_ref if c.isdigit()))
    col = column_index_from_string(col_letter)
    return (row, col)


def copy_template_to(target_path: str, template_path: str) -> None:
    """Copy template workbook to target."""
    Path(target_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, target_path)


def write_cells(workbook_path: str, updates: dict[str, str | bool | int | float]) -> None:
    """
    Write values to cells. updates: { "cell_ref": value } or { "sheet!cell": value }
    Value types: bool -> True/False, int/float -> number, str -> string
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    for key, value in updates.items():
        if "!" in key:
            sheet_name, cell_ref = key.split("!", 1)
            ws = wb[sheet_name.strip()]
        else:
            cell_ref = key
            ws = wb.active
        row, col = _cell_ref_to_tuple(cell_ref.strip())
        if isinstance(value, bool):
            ws.cell(row=row, column=col, value=value)
        elif isinstance(value, (int, float)):
            ws.cell(row=row, column=col, value=value)
        elif value is None or value == "":
            ws.cell(row=row, column=col, value=None)
        else:
            ws.cell(row=row, column=col, value=value)
    wb.save(workbook_path)


def _form_value_to_cell(value) -> str | bool | int | float | None:
    """Convert form JSON value to Excel cell value."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        if value.lower() in ("true", "igen", "yes"):
            return True
        if value.lower() in ("false", "nem", "no"):
            return False
        if value.lower() in ("nem releváns", "pályázatból szerzi be", "már rendelkezik vele"):
            return value  # Preserve original casing
        try:
            return int(value)
        except ValueError:
            pass
        try:
            return float(value)
        except ValueError:
            pass
        return value
    return str(value)


def apply_form_data(workbook_path: str, form_data: dict, cell_map: dict) -> None:
    """Apply form_data (key -> value) to workbook using cell_map {key: 'Sheet!Cell'}."""
    updates = {}
    for key, value in form_data.items():
        if key in cell_map:
            cell_ref = cell_map[key]
            cv = _form_value_to_cell(value)
            if cv is not None:
                updates[cell_ref] = cv
            elif key.startswith("sw_mennyiseg_") and value == "":
                updates[cell_ref] = "nem releváns"
    if updates:
        write_cells(workbook_path, updates)


def read_budget_preview(workbook_path: str) -> dict:
    """Read budget summary from Költségvetés sheet (F oszlop = nettó). Brutto = netto * 1.27."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["Költségvetés"]
    cells = {
        "hardver_osszesen": ("F", 4),
        "szoftver_osszesen": ("F", 6),
        "kepzes_osszesen": ("F", 8),
        "szolgaltatasok_osszesen": ("F", 10),
        "egyeb_koltsegek": ("F", 12),
        "projektkoltseg_osszesen": ("F", 14),
        "onero": ("F", 16),
        "tamogatas_osszesen": ("F", 18),
    }
    result = {}
    for k, (col, row) in cells.items():
        val = ws.cell(row=row, column=column_index_from_string(col)).value
        num = round(val, 0) if isinstance(val, (int, float)) else None
        result[k] = num
        if num is not None and k in ("projektkoltseg_osszesen", "onero", "tamogatas_osszesen"):
            result[f"{k}_brutto"] = round(num * 1.27, 0)
    return result


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="cmd")
    copy_p = sub.add_parser("copy")
    copy_p.add_argument("--template", required=True)
    copy_p.add_argument("--target", required=True)

    write_p = sub.add_parser("write")
    write_p.add_argument("--workbook", required=True)
    write_p.add_argument("--updates", required=True, help="JSON object of cell->value")

    apply_p = sub.add_parser("apply")
    apply_p.add_argument("--workbook", required=True)
    apply_p.add_argument("--form-data-file", required=True, help="Path to JSON file with field key->value")
    apply_p.add_argument("--cell-map-file", required=True, help="Path to JSON file with key->Sheet!Cell")

    preview_p = sub.add_parser("preview")
    preview_p.add_argument("--workbook", required=True)

    args = parser.parse_args()
    if args.cmd == "copy":
        copy_template_to(args.target, args.template)
    elif args.cmd == "write":
        updates = json.loads(args.updates)
        write_cells(args.workbook, updates)
    elif args.cmd == "apply":
        with open(args.form_data_file) as f:
            form_data = json.load(f)
        with open(args.cell_map_file) as f:
            cell_map = json.load(f)
        apply_form_data(args.workbook, form_data, cell_map)
    elif args.cmd == "preview":
        print(json.dumps(read_budget_preview(args.workbook)))


if __name__ == "__main__":
    main()
